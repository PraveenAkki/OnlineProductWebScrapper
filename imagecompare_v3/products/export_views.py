"""
products/export_views.py
--------------------------
Excel export endpoints.

Placement:  products/export_views.py   (new file — create it)

Add to products/urls.py (3 lines — see bottom of this file):
    from .export_views import ExportSearchView, ExportSearchLensOnlyView, ExportAllSearchesView

    path("searches/<int:pk>/export/",      ExportSearchView.as_view(),        name="export-search"),
    path("searches/<int:pk>/export/lens/", ExportSearchLensOnlyView.as_view(), name="export-lens"),
    path("export/all/",                    ExportAllSearchesView.as_view(),    name="export-all"),

Install:  python -m pip install openpyxl==3.1.2

Works on Render.com — streams from memory (io.BytesIO), no disk write.
"""

import io
import re
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status

from .models import SearchHistory, GoogleLensResult, Product


# ── Styles ────────────────────────────────────────────────────────────────────

HEADER_FONT  = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL  = PatternFill("solid", fgColor="1E3A5F")   # dark navy
ALT_FILL     = PatternFill("solid", fgColor="F2F6FF")   # light blue-grey
GREEN_FILL   = PatternFill("solid", fgColor="D6F5E3")   # lowest price
RED_FILL     = PatternFill("solid", fgColor="FFE5E5")   # highest price
BORDER_SIDE  = Side(style="thin", color="CCCCCC")
CELL_BORDER  = Border(left=BORDER_SIDE, right=BORDER_SIDE,
                      top=BORDER_SIDE,  bottom=BORDER_SIDE)
CENTER       = Alignment(horizontal="center", vertical="center")
LINK_FONT    = Font(color="1D4ED8", underline="single")


def _header(ws, row=1):
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = CELL_BORDER


def _col_widths(ws, widths: list):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _hyperlink(cell, url):
    if url:
        cell.hyperlink = url
        cell.font = LINK_FONT


# ── Sheet builders ────────────────────────────────────────────────────────────

def _sheet_summary(wb, search, product_count, lens_count):
    ws = wb.create_sheet("Summary", index=0)
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 45

    # Title
    ws["A1"] = "ShopLens — Search Export"
    ws["A1"].font = Font(bold=True, size=14, color="1E3A5F")
    ws.merge_cells("A1:B1")
    ws.append([])

    shopping_count = GoogleLensResult.objects.filter(search=search, result_type="shopping").count()
    visual_count   = GoogleLensResult.objects.filter(search=search, result_type="visual").count()
    scraped_count  = GoogleLensResult.objects.filter(search=search, scraped=True).count()
    priced_count   = Product.objects.filter(search=search, price_numeric__gt=0).count()
    prices         = list(Product.objects.filter(
                         search=search, price_numeric__gt=0
                     ).values_list("price_numeric", flat=True))

    rows = [
        ("Search ID",                  search.id),
        ("Keyword",                    search.search_keyword or "—"),
        ("Category",                   search.category or "—"),
        ("Classifier Phase",           search.classifier_phase or "—"),
        ("Detected Label",             search.detected_label or "—"),
        ("Detected Color",             search.detected_color or "—"),
        ("Confidence",                 f"{search.confidence:.0%}" if search.confidence else "—"),
        ("Exported At",                datetime.now().strftime("%Y-%m-%d %H:%M")),
        ("", ""),
        ("Total Lens Results",         lens_count),
        ("Shopping (with price)",      shopping_count),
        ("Visual (no price)",          visual_count),
        ("Scraped",                    scraped_count),
        ("Products in DB",             product_count),
        ("Products with Price",        priced_count),
        ("Lowest Price (₹)",           min(prices) if prices else "—"),
        ("Highest Price (₹)",          max(prices) if prices else "—"),
        ("Average Price (₹)",          round(sum(prices) / len(prices)) if prices else "—"),
    ]
    for label, value in rows:
        r = ws.max_row + 1
        ws.cell(r, 1, label).font = Font(bold=True, color="333333")
        ws.cell(r, 2, value)

    return ws


def _sheet_products(wb, search, sheet_name="Products"):
    ws = wb.create_sheet(sheet_name[:31])
    headers = [
        "#", "Website", "Product Name", "Price (Display)", "Price (₹)",
        "Rating", "Reviews", "Delivery", "Discount / Tag",
        "Product Link", "Image URL", "Type", "Exported At",
    ]
    ws.append(headers)
    _header(ws)
    ws.freeze_panes = "A2"

    products = Product.objects.filter(search=search).order_by("price_numeric")
    if not products.exists():
        ws.append(["No products found for this search."])
        return ws, 0

    prices    = [p.price_numeric for p in products if p.price_numeric > 0]
    min_price = min(prices) if prices else None
    max_price = max(prices) if prices else None

    for i, p in enumerate(products, 1):
        # Determine type
        prod_type = "Shopping" if p.price_numeric > 0 else "Visual"

        ws.append([
            i,
            p.website or "",
            p.product_name or "",
            p.price or "",
            p.price_numeric or "",
            p.rating or "",
            p.reviews or "",
            p.delivery or "",
            p.discount or "",
            p.product_link or "",
            p.product_image or "",
            prod_type,
            datetime.now().strftime("%Y-%m-%d"),
        ])

        excel_row = i + 1
        if min_price and p.price_numeric == min_price and p.price_numeric > 0:
            for c in ws[excel_row]: c.fill = GREEN_FILL
        elif max_price and p.price_numeric == max_price and p.price_numeric > 0 and len(prices) > 1:
            for c in ws[excel_row]: c.fill = RED_FILL
        elif i % 2 == 0:
            for c in ws[excel_row]: c.fill = ALT_FILL

        # Clickable product link (column 10)
        _hyperlink(ws.cell(excel_row, 10), p.product_link)

    _col_widths(ws, [4, 14, 50, 14, 12, 8, 10, 22, 20, 45, 40, 10, 14])
    return ws, products.count()


def _sheet_lens(wb, search, sheet_name="Lens Results"):
    ws = wb.create_sheet(sheet_name[:31])
    headers = [
        "#", "Type", "Rank", "Title", "Source",
        "Price (Google)", "Price (₹)", "Rating", "Reviews",
        "Delivery", "Tag", "Scraped?", "Scraped Price",
        "Link", "Thumbnail", "Image URL", "Saved At",
    ]
    ws.append(headers)
    _header(ws)
    ws.freeze_panes = "A2"

    rows = GoogleLensResult.objects.filter(search=search).order_by("result_type", "rank")
    for i, r in enumerate(rows, 1):
        ws.append([
            i,
            r.result_type.capitalize(),
            r.rank,
            r.title or "",
            r.source or "",
            r.price or "",
            r.price_numeric if r.price_numeric else "",
            r.rating or "",
            r.reviews or "",
            r.delivery or "",
            r.tag or "",
            "Yes" if r.scraped else "No",
            r.scraped_price or "",
            r.link or "",
            r.thumbnail or "",
            r.image_url or "",
            r.created_at.strftime("%Y-%m-%d %H:%M") if r.created_at else "",
        ])
        excel_row = i + 1
        if r.result_type == "shopping" and r.price_numeric > 0:
            for c in ws[excel_row]: c.fill = GREEN_FILL
        elif i % 2 == 0:
            for c in ws[excel_row]: c.fill = ALT_FILL

        # Clickable link (column 14)
        _hyperlink(ws.cell(excel_row, 14), r.link)

    _col_widths(ws, [4, 12, 6, 50, 20, 14, 10, 8, 10, 20, 15, 10, 14, 45, 35, 35, 16])
    return ws, rows.count()


def _make_workbook(search):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    product_count = Product.objects.filter(search=search).count()
    lens_count    = GoogleLensResult.objects.filter(search=search).count()

    _sheet_summary(wb, search, product_count, lens_count)
    _sheet_products(wb, search)
    _sheet_lens(wb, search)
    return wb


def _xlsx_response(wb, filename: str) -> HttpResponse:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(
        buf.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    resp["Content-Disposition"] = f'attachment; filename="{filename}"'
    resp["Access-Control-Expose-Headers"] = "Content-Disposition"
    return resp


def _slug(text, max_len=30):
    return re.sub(r"[^\w\s-]", "", text or "search")[:max_len].strip().replace(" ", "_")


# ── Views ──────────────────────────────────────────────────────────────────────

class ExportSearchView(APIView):
    """
    GET /api/searches/<id>/export/
    Download full xlsx: Summary + Products + Lens Results for one search.
    Green rows = lowest price. Red rows = highest price.
    """
    def get(self, request, pk):
        try:
            search = SearchHistory.objects.get(pk=pk)
        except SearchHistory.DoesNotExist:
            return Response({"error": f"Search id={pk} not found."}, status=status.HTTP_404_NOT_FOUND)

        wb = _make_workbook(search)
        return _xlsx_response(wb, f"shoplens_search_{pk}_{_slug(search.search_keyword)}.xlsx")


class ExportSearchLensOnlyView(APIView):
    """
    GET /api/searches/<id>/export/lens/
    Download only raw lens results (all links, prices, thumbnails).
    Useful when you want to see every URL Google found — with and without price.
    """
    def get(self, request, pk):
        try:
            search = SearchHistory.objects.get(pk=pk)
        except SearchHistory.DoesNotExist:
            return Response({"error": f"Search id={pk} not found."}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        lens_count = GoogleLensResult.objects.filter(search=search).count()
        _sheet_summary(wb, search, 0, lens_count)
        _sheet_lens(wb, search)
        return _xlsx_response(wb, f"shoplens_lens_{pk}_{_slug(search.search_keyword)}.xlsx")


class ExportAllSearchesView(APIView):
    """
    GET /api/export/all/
    Download all searches in one workbook.
    Sheet 1 = master summary table of all searches.
    Sheets 2–N = one Products sheet per search (up to 15 searches).
    """
    def get(self, request):
        searches = SearchHistory.objects.all().order_by("-created_at")[:15]
        if not searches:
            return Response({"error": "No searches found."}, status=status.HTTP_404_NOT_FOUND)

        wb = openpyxl.Workbook()
        wb.remove(wb.active)

        # Master summary
        ws = wb.create_sheet("All Searches")
        ws.append(["#", "Search ID", "Keyword", "Category", "Phase",
                   "Products", "Priced", "Shopping", "Visual", "Created"])
        _header(ws)
        ws.freeze_panes = "A2"

        for i, search in enumerate(searches, 1):
            product_count  = Product.objects.filter(search=search).count()
            priced_count   = Product.objects.filter(search=search, price_numeric__gt=0).count()
            shopping_count = GoogleLensResult.objects.filter(search=search, result_type="shopping").count()
            visual_count   = GoogleLensResult.objects.filter(search=search, result_type="visual").count()

            ws.append([
                i, search.id,
                search.search_keyword or "—",
                search.category or "—",
                search.classifier_phase or "—",
                product_count, priced_count,
                shopping_count, visual_count,
                search.created_at.strftime("%Y-%m-%d %H:%M") if search.created_at else "—",
            ])
            if i % 2 == 0:
                for c in ws[i + 1]: c.fill = ALT_FILL

            # One products sheet per search
            sheet_name = f"#{search.id} {(_slug(search.search_keyword) or 'search')[:18]}"
            _sheet_products(wb, search, sheet_name=sheet_name[:31])

        _col_widths(ws, [4, 10, 40, 14, 16, 10, 10, 12, 12, 18])
        return _xlsx_response(wb, f"shoplens_all_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
